import xlsxwriter 
import openpyxl
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter.ttk import *
from tkinter import *

from tkinter import messagebox
from tkinter import filedialog as fd
from tkinter.filedialog import asksaveasfile
import time 

root = tk.Tk()
root.title("Create Read & Delete")
w, h = root.winfo_screenwidth(), root.winfo_screenheight()
root.geometry("%dx%d+0+0" % (w, h))
root.resizable(True, True)
root.configure(bg='#0680f9')
root.iconphoto(False, tk.PhotoImage(file='crud.png'))
bg = PhotoImage(file = "tech.png") 

 
# Create Canvas 
canvas1 = Canvas( root, width = 400, height = 400) 
  
canvas1.pack(fill = "both", expand = True) 
  
# Display image 
canvas1.create_image( 0, 0, image = bg,  anchor = "nw")

logo = tk.PhotoImage(file = "logo.png")
tk.Label(canvas1, image=logo,bg="#0680f9").pack() 
global path
path=''
text=Text(root)

label1=Label(root,text="Key",width=20,height=2,bg="#0680f9")
label2=Label(root,text="Value",width=20,height=2,bg="#0680f9")
label3=Label(root,text="Time-To-Live",width=20,height=2,bg="#0680f9")

e1=Entry(root,width=20,borderwidth=8)
# e1.grid(row=0,column=1)
e2=Entry(root,width=20,borderwidth=8)
# e2.grid(row=1,column=1)
e3=Entry(root,width=20,borderwidth=8)

label1_canvas = canvas1.create_window( 10, 400,  anchor = "nw", window = label1)
e1_canvas = canvas1.create_window( 200, 400,  anchor = "nw", window = e1)
label2_canvas = canvas1.create_window( 400, 400,  anchor = "nw", window = label2)
e2_canvas = canvas1.create_window( 600, 400,  anchor = "nw", window = e2)
label3_canvas = canvas1.create_window( 10, 500,  anchor = "nw", window = label3)
e3_canvas = canvas1.create_window( 200, 500,  anchor = "nw", window = e3)

def Open():
    global path
    filetype = [("excel file", "*.xlsx")]
    path = fd.askopenfilename(title="Choose a excel file", filetypes=filetype)
    df = pd.read_excel(path)
    data = pd.read_excel(path, index_col ="Key" )
    for i in range(len(df)):#deleting the records whose time to live property has expired
        key=df.loc[i, "Key"]
        value=df.loc[i, "Value"]
        ti=df.loc[i,"Creation Time"]
        timestamp=df.loc[i,"Timestamp"]
        delta = (time.time()-float(ti))
        if(delta>float(timestamp)):
            data.drop([key], inplace = True)
    data.to_excel(path)




def Save():
    global path
    path = fd.asksaveasfilename(confirmoverwrite=False, defaultextension=".xlsx")
    if(len(path)==0):
        path='argho.xlsx'
        workbook = xlsxwriter.Workbook(path) 
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0,"Key")
        worksheet.write(0,1,"Value")
        worksheet.write(0,2,"Timestamp")
        worksheet.write(0,3,"Creation Time")
        workbook.close()
    else:
        workbook = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0,"Key")
        worksheet.write(0,1,"Value")
        workbook.close()
    

def Register():
    global path
    if(path==''):
        messagebox.showerror("Error", "Select a file")
        return
    key = e1.get()
    value = e2.get()
    timestamp = e3.get()
    if(key=='' or value==''):
        messagebox.showerror("Error", "Fill key value")
        return
    df = pd.read_excel(path)
    ans = df['Key'].where(df['Key']==key).dropna()
    if(not ans.empty):#checking for duplicate key
    	messagebox.showerror("Error", "The key already exists..please try out a unique key") 
    	return
    if(timestamp==''):
        timestamp = time.time()
    key_value = (key,value,float(timestamp),time.time())
    wb = openpyxl.load_workbook(path)#loading the current workbook
    sheet = wb.active
    sheet.append(key_value)#insert the key and value pair
    wb.save(path)
    messagebox.showinfo("Success", "Data inserted successfully")
    wb.close()

def ShowRecord():
    global path
    if(path==''):
        messagebox.showerror("Error", "Select a file")
        return
    key=e1.get()
    if(key==''):
        messagebox.showerror("Error", "Provide key for the desired value")
        return
    df = pd.read_excel(path)
    ans = df['Value'].where(df['Key']==key).dropna()
    if(ans.empty):
        messagebox.showerror("Error", "Record not found")
        return
    else:
        timestamp = df['Timestamp'].where(df['Key']==key).dropna()
        ti = df['Creation Time'].where(df['Key']==key).dropna()
        delta = (time.time()-float(ti.values[0]))
        if(delta>float(timestamp.values[0])):
            messagebox.showerror("Error", "Record has expired")
            return
    e2.delete(0, END)
    e3.delete(0, END)
    e2.insert(0,ans.values[0])#if the key is available showing its value

def Delete():
    global path
    if(path==''):
        messagebox.showerror("Error", "Select a file")
        return
    key=e1.get()
    if(key==''):
        messagebox.showerror("Error", "Provide key for which the data is to be deleted")
        return
    df = pd.read_excel(path)
    data = pd.read_excel(path, index_col ="Key" )
    ans = df['Value'].where(df['Key']==key).dropna()#creating a series for the specified key while droping the unwanted values  
    if(not ans.empty):
        timestamp = df['Timestamp'].where(df['Key']==key).dropna()
        ti = df['Creation Time'].where(df['Key']==key).dropna()
        delta = (time.time()-float(ti.values[0]))
        if(delta>float(timestamp.values[0])):
            messagebox.showerror("Error", "Record has expired")
            return
        data.drop([key], inplace = True)#deleting the specified key
        data.to_excel(path)#saving the current workbook
        messagebox.showinfo("Success", "Record Deleted") 
    else: 
    	messagebox.showerror("Error", "Record is not present")
    	
def Showall():
    global path
    if(path==''):
        messagebox.showerror("Error", "Select a file")
        return
    df = pd.read_excel(path)
    data = pd.read_excel(path, index_col ="Key" )
    for i in range(len(df)):#deleting the files whose time to live property has expired
        key=df.loc[i, "Key"]
        value=df.loc[i, "Value"]
        ti=df.loc[i,"Creation Time"]
        timestamp=df.loc[i,"Timestamp"]
        delta = (time.time()-float(ti))
        if(delta>float(timestamp)):
            data.drop([key], inplace = True)
    data.to_excel(path)#saving the current workbook
    class A(Frame):#creating a table to show all the values
        def __init__(self, parent):
            Frame.__init__(self, parent)
            self.CreateUI()
            self.LoadTable()
            self.grid(sticky=(N, S, W, E))
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
        def CreateUI(self):
            tv= Treeview(self)
            tv['columns']=('Key', 'Value')
            tv.heading('#0',text='Key',anchor='center')
            tv.column('#0',anchor='center')
            tv.heading('#1', text='Value', anchor='center')
            tv.column('#1', anchor='center')
            tv.grid(sticky=(N,S,W,E))
            self.treeview = tv
            self.grid_rowconfigure(0,weight=1)
            self.grid_columnconfigure(0,weight=1)
        def LoadTable(self):
        	df = pd.read_excel(path)#reading the file from the workbook
        	key=""
        	value=""
        	for i in range(len(df)):#inserting all the values in table one by one
        		key=df.loc[i, "Key"]
        		value=df.loc[i, "Value"]
        		self.treeview.insert("",'end',text=key,values=value)
    root=Tk()
    root.title("Showing All Records")
    A(root)
def Clear():
    e1.delete(0, END)
    e2.delete(0, END)
    e3.delete(0, END)

button1=Button(root,text="Create",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Register)
button2=Button(root,text="Delete",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Delete)
button4=Button(root,text="Read",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=ShowRecord)
button5=Button(root,text="Show",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Showall)
button6=Button(root,text="Clear",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Clear)
button7=Button(root,text="Open",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Open)
button8=Button(root,text="Initialize",width=5,height=2,font="Arial 20", bg="#000000", fg="white", 
        pady=-2, bd=0, highlightthickness=0, activebackground="#000000", activeforeground="red",command=Save)

button1_canvas = canvas1.create_window( 1060, 400,  anchor = "nw", window = button1) 
  
button2_canvas = canvas1.create_window( 1320, 400, anchor = "nw", window = button2) 
  
button4_canvas = canvas1.create_window( 1190, 400,  anchor = "nw", window = button4) 
  
button5_canvas = canvas1.create_window( 1450, 400, anchor = "nw", window = button5) 
  
button6_canvas = canvas1.create_window( 1580, 400, anchor = "nw", window = button6)
button7_canvas = canvas1.create_window( 800, 400,  anchor = "nw", window = button7) 
  
button8_canvas = canvas1.create_window( 930, 400, anchor = "nw", window = button8) 
  

root.mainloop()